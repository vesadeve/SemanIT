# Script for similarity search from other documents
from gensim import corpora, models, similarities
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
import snowballstemmer
import os

# Place index-file, dictionary and mm-file to subdirectory \index
script_dir = os.path.dirname(__file__) #<-- absolute dir the script is in
#script_dir = ''

# Model creation
dictionary = corpora.Dictionary.load(os.path.join(script_dir, 'Index/ErrorReportStrip_v01_IT.dict'))
corpus = corpora.MmCorpus((os.path.join(script_dir, 'Index/ErrorReportStrip_v01_IT.mm'))) # corpus created from previous error data
lsi = models.LsiModel(corpus, id2word=dictionary, num_topics=200) # this creates a model that is being used

# Use italian stemmer
stemmer = snowballstemmer.stemmer('italian')

def strip(s):
    s = re.sub(r'[^\w\s]',' ',s) # remove non-word characters
    s = re.sub(r'[\d+]','',s) # remove numbers
    stoplist = set('''non destra dx sinistra sx era è mancante mancanti manca rotto rotta danneggiato lento 
          lenta allentata allentato anche ancora lavorano lavoro lavorato fa fai ancora apre chiude chiudono chiusura
          aperto chiuso e sistemare controllare montare d di da in con per su tra fra la lo il una uno un
          1 2 3 4 5 6 7 8 9 0 ad al agli agl alla alle con col coi da dal dallo dai dagli dall dagl dalla 
          dalle di del dello dei degli dell degl della delle in nel nello nei 
          negli nell negl nella nelle su sul sullo sui sugli sull sugl sulla sulle per tra contro 
          io tu lui lei noi voi loro mio mia miei mie tuo tua tuoi tue suo sua suoi sue nostro 
          nostra nostri nostre vostro vostra vostri vostre mi ti ci vi lo la li le gli ne il un uno una 
          ma ed se perché anche come dov dove che chi cui non più quale quanto quanti 
          quanta quante quello quelli quella quelle questo questi questa queste si tutto tutti 
          a c e i l o ho hai ha abbiamo avete hanno abbia abbiate abbiano 
          avrò avrai avrà avremo avrete avranno avrei avresti avrebbe avremmo avreste avrebbero 
          avevo avevi aveva avevamo avevate avevano ebbi avesti ebbe avemmo aveste ebbero avessi 
          avesse avessimo avessero avendo avuto avuta avuti avute sono sei è siamo 
          siete sia siate siano sarò sarai sarà saremo sarete saranno sarei saresti sarebbe saremmo 
          sareste sarebbero ero eri era eravamo eravate erano fui fosti fu fummo foste furono 
          fossi fosse fossimo fossero essendo faccio fai facciamo fanno faccia facciate 
          facciano farò farai farà faremo farete faranno farei faresti farebbe faremmo 
          fareste farebbero facevo facevi faceva facevamo facevate facevano feci facesti 
          fece facemmo faceste fecero facessi facesse facessimo facessero facendo sto stai sta 
          stiamo stanno stia stiate stiano starò starai starà staremo starete staranno starei 
          staresti starebbe staremmo stareste starebbero stavo stavi stava stavamo stavate 
          stavano stetti stesti stette stemmo steste stettero stessi stesse stessimo stessero stando'''.split())
    s = [word for word in s.lower().split() if word not in stoplist]
    return s

# load error report to be analyzed
wc=load_workbook('ErrorReport_XLS.xlsx')
ws=wc.active

print("Reading error reports (notes)")
defects = []
for row in range(2, ws.max_row+1): # here we are reading the whole error report in memory - if needed could be done line by line
    defects.append({
    'NUM': ws['A' + str(row)].value,
    'KEY PIANO': ws['B' + str(row)].value,
    'TELAIO': ws['C' + str(row)].value,
    'MODELLO': ws['D' + str(row)].value,
    'DT FINE COLLAUDO': ws['E' + str(row)].value,
    'PTZ': ws['F' + str(row)].value,
    'GRUPPO': ws['G' + str(row)].value,
    'SOTTOGRUPPO': ws['H' + str(row)].value,
    'DETTAGLIO NC': ws['I' + str(row)].value,
    'NOTE': ws['J' + str(row)].value})
#    'DT USCITA LINEA': ws['K' + str(row)].value,
#    'DATA NC': ws['L' + str(row)].value,
#    'STATO': ws['M' + str(row)].value,
#    'TIPO NC': ws['N' + str(row)].value,
#    'RAGGRUPPAMENTO NC': ws['O' + str(row)].value,
#    'TEMPO RIPARAZIONE': ws['P' + str(row)].value,
#    'REP RESP': ws['Q' + str(row)].value,
#    'VAL DEMERITO': ws['R' + str(row)].value,
#    'ARTICOLO': ws['S' + str(row)].value,
#    'LUNGHEZZA': ws['T' + str(row)].value,
#    'CORREZIONE': ws['U' + str(row)].value,
#    'AZIONI CORRECTIVE': ws['V' + str(row)].value,
#    'REP AC': ws['W' + str(row)].value,
#    'DATA ATTUAZIONE': ws['X' + str(row)].value,
#    'NC riparabile in linea di mont.': ws['Y' + str(row)].value,
#    'VER EFF': ws['Z' + str(row)].value})

documents=[] # list of comment fields for analysis
for row in range(2, ws.max_row+1):
    documents.append(ws['J' + str(row)].value) # this is the comment field we are using for analysis

# remove common words and tokenize
print("Removing common words")
texts =[]
for document in documents:
    if document: # only strip non-empty documents
        texts.append(strip(document))
    else:
        texts.append('')

# stem the texts (use snowball to stem the italian words)
print("Stemming the notes")
for text,i in zip(texts,range(0,len(texts))):
     texts[i] = stemmer.stemWords(text)

print("Creating dictonary of notes")
errors_to_analyze = [dictionary.doc2bow(text) for text in texts]

# Add here another corpus where to look for "doc" string similarity
# conversions - use lsi defined earlier to transform this new document collection (corpus) to same
#   LSI space as earlier training corpus - lets call this error_reports

print("Transform to LSI space")
index = similarities.MatrixSimilarity(lsi[errors_to_analyze]) # transform error reports to LSI space and index it
index.save((os.path.join(script_dir, 'Tmp/Errors_to_analyze.index')))
index = similarities.MatrixSimilarity.load((os.path.join(script_dir, 'Tmp/Errors_to_analyze.index')))

# Here is the phrase (document) we are looking for
# Read search phrases line by line from excel
# create a new worksheet for each search phrase results including first 50 results
# load error report to be analyzed
wb=load_workbook('ErrorsToSearch.xlsx')
wbs=wb.active

search_strings=[]
tasto = []
for row in range(2, wbs.max_row+1):
    if wbs['A' + str(row)].value:
        search_strings.append(wbs['A' + str(row)].value)
        tasto.append(wbs['B'+ str(row)].value)
    else: break

# remove common words and tokenize
print("Reading and stemming search keys (tasto)")
strings =[]
for document,i in zip(search_strings, range(0,len(search_strings))):
    print(tasto[i], i+1, document)
    strings.append(strip(document))

# stem the texts (use snowball to stem the italian words)
for string,i in zip(strings,range(0,len(strings))):
     strings[i] = stemmer.stemWords(string)

print("Doing similarity search and scoring")
for doc,i in zip(strings,range(0,len(strings))):
    vec_bow = dictionary.doc2bow(doc) # create a vector from document (using dictionary created earlier)
    vec_lsi = lsi[vec_bow] # convert the query to LSI space
    sims = index[vec_lsi] # perform a similarity query against the corpus
    sims = sorted(enumerate(sims), key=lambda item: -item[1])

    wbs = wb.create_sheet(tasto[i])
    column_width = (12, 12, 22, 12, 15, 22, 22, 22, 22, 15, 40) # print score as well
    columns = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K')
    for col,width in zip(columns, column_width):
        wbs.column_dimensions[col].width = width
    wbs.freeze_panes = 'A3'
    wbs['A1'] = ' '.join(doc)  # join doc to  print search terms
    wbs['A2'] = 'NUM' # 77pt
    wbs['B2'] = 'KEY PIANO' # 112 pt
    wbs['C2'] = 'TELAIO' # 77
    wbs['D2'] = 'MODELLO' # 77
    wbs['E2'] = 'DT FINE COLLAUDO' #  100
    wbs['F2'] = 'PTZ' # 100
    wbs['G2'] = 'GRUPPO' # 100
    wbs['H2'] = 'SOTTOGRUPPO' # 100
    wbs['I2'] = 'DETTAGLIO NC' # 77
    wbs['J2'] = 'SCORE' # 77
    wbs['K2'] = 'NOTE' # 77

    for sim in range(0, 100):
        # read the line from error excel and copy to analysis excel 180225 add score as well
        wbs['A' + str(sim + 3)].value = defects[sims[sim][0]]['NUM']
        wbs['B' + str(sim + 3)].value = defects[sims[sim][0]]['KEY PIANO']
        wbs['C' + str(sim + 3)].value = defects[sims[sim][0]]['TELAIO']
        wbs['D' + str(sim + 3)].value = defects[sims[sim][0]]['MODELLO']
        wbs['E' + str(sim + 3)].value = defects[sims[sim][0]]['DT FINE COLLAUDO']
        wbs['F' + str(sim + 3)].value = defects[sims[sim][0]]['PTZ']
        wbs['G' + str(sim + 3)].value = defects[sims[sim][0]]['GRUPPO']
        wbs['H' + str(sim + 3)].value = defects[sims[sim][0]]['SOTTOGRUPPO']
        wbs['I' + str(sim + 3)].value = defects[sims[sim][0]]['DETTAGLIO NC']
        wbs['J' + str(sim + 3)].value = sims[sim][1] # similarity score
        wbs['K' + str(sim + 3)].alignment = Alignment(wrapText=True)
        wbs['K' + str(sim + 3)].value = defects[sims[sim][0]]['NOTE']


print("Saving results excel")
wb.save("ErrorAnalysis.xlsx")
print("Done")
